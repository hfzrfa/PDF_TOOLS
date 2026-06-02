"""
PDF Compressor  –  by hfzrfa
Drag & Drop  ·  Custom output  ·  Multiple compression levels
"""

import os
import sys
import json
import re
import shutil
import subprocess
import threading
from pathlib import Path
from tkinter import filedialog, messagebox

import customtkinter as ctk

try:
    import fitz  # PyMuPDF
except ImportError:
    print("PyMuPDF not installed. Run: pip install PyMuPDF")
    sys.exit(1)

# ── Drag-and-Drop ────────────────────────────────────────────────────
DND_AVAILABLE = False
try:
    from tkinterdnd2 import TkinterDnD, DND_FILES

    class _RootBase(ctk.CTk, TkinterDnD.DnDWrapper):
        def __init__(self, *args, **kwargs):
            super().__init__(*args, **kwargs)
            self.TkdndVersion = TkinterDnD._require(self)

    DND_AVAILABLE = True
except Exception:
    _RootBase = ctk.CTk


# ── Helpers ──────────────────────────────────────────────────────────
def fmt(size: int) -> str:
    if size < 1024:
        return f"{size} B"
    if size < 1024 * 1024:
        return f"{size / 1024:.1f} KB"
    return f"{size / (1024 * 1024):.2f} MB"


def compress_single_pdf(input_path, output_path, quality="medium", callback=None):
    original_size = os.path.getsize(input_path)
    doc = fitz.open(input_path)
    total_pages = len(doc)
    seen_images = set()

    jpeg_q = {"low": 25, "medium": 55, "high": 82}[quality]
    garbage = {"low": 4, "medium": 3, "high": 2}[quality]

    # ── Try to recompress images (best-effort) ───────────────
    for idx in range(total_pages):
        if callback:
            callback(int((idx / total_pages) * 80),
                     f"Page {idx + 1}/{total_pages}")
        try:
            page = doc[idx]
            for img in page.get_images(full=True):
                xref = img[0]
                smask = img[1]
                if xref in seen_images or smask:
                    continue
                try:
                    pix = fitz.Pixmap(doc, xref)
                    if pix.alpha or pix.width < 64 or pix.height < 64:
                        continue
                    # Convert CMYK or other colorspaces to RGB
                    if pix.n > 3:
                        pix = fitz.Pixmap(fitz.csRGB, pix)
                    new_bytes = pix.tobytes("jpeg", jpg_quality=jpeg_q)
                    old_bytes = doc.extract_image(xref).get("image", b"")
                    if old_bytes and len(new_bytes) < len(old_bytes):
                        page.replace_image(xref, stream=new_bytes)
                        seen_images.add(xref)
                except Exception:
                    continue
        except Exception:
            continue

    if callback:
        callback(90, "Saving...")

    # ── Save with compatible parameters ──────────────────────
    try:
        try:
            doc.save(output_path, garbage=garbage, deflate=True, clean=True)
        except TypeError:
            # Fallback for older PyMuPDF versions
            doc.save(output_path, garbage=garbage, deflate=True)
    finally:
        doc.close()

    if callback:
        callback(100, "Done")
    return original_size, os.path.getsize(output_path)


def unique_output_path(output_dir, base_name, extension):
    path = os.path.join(output_dir, f"{base_name}{extension}")
    if not os.path.exists(path):
        return path

    index = 2
    while True:
        candidate = os.path.join(output_dir, f"{base_name}_{index}{extension}")
        if not os.path.exists(candidate):
            return candidate
        index += 1


def convert_pdf_to_word(input_path, output_path, callback=None):
    if not hasattr(fitz.Rect, "get_area"):
        fitz.Rect.get_area = lambda rect: max(0, rect.width) * max(0, rect.height)

    try:
        from pdf2docx import Converter
    except ImportError as exc:
        raise RuntimeError("Install pdf2docx first: pip install pdf2docx") from exc

    if callback:
        callback(10, "Reading PDF...")

    converter = Converter(input_path)
    try:
        if callback:
            callback(35, "Converting pages...")
        converter.convert(output_path, start=0, end=None)
    finally:
        converter.close()

    if callback:
        callback(100, "Done")


def _find_soffice():
    candidate = shutil.which("soffice") or shutil.which("libreoffice")
    if candidate:
        return candidate

    for path in (
        r"C:\Program Files\LibreOffice\program\soffice.exe",
        r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
    ):
        if os.path.exists(path):
            return path
    return None


def _convert_with_soffice(input_path, output_path):
    soffice = _find_soffice()
    if not soffice:
        raise RuntimeError("Microsoft Office or LibreOffice is required.")

    output_dir = os.path.dirname(output_path)
    command = [
        soffice,
        "--headless",
        "--convert-to",
        "pdf",
        "--outdir",
        output_dir,
        input_path,
    ]
    completed = subprocess.run(
        command, capture_output=True, text=True, timeout=180)
    if completed.returncode != 0:
        msg = (completed.stderr or completed.stdout or "LibreOffice failed.")
        raise RuntimeError(msg.strip())

    produced = os.path.join(output_dir, f"{Path(input_path).stem}.pdf")
    if not os.path.exists(produced):
        pdfs = sorted(
            Path(output_dir).glob("*.pdf"),
            key=lambda p: p.stat().st_mtime,
            reverse=True,
        )
        produced = str(pdfs[0]) if pdfs else produced

    if not os.path.exists(produced):
        raise RuntimeError("PDF output was not created.")

    if os.path.normcase(os.path.abspath(produced)) != os.path.normcase(
            os.path.abspath(output_path)):
        if os.path.exists(output_path):
            os.remove(output_path)
        os.replace(produced, output_path)


def convert_word_to_pdf(input_path, output_path, callback=None):
    if callback:
        callback(15, "Opening Word...")

    try:
        import pythoncom
        import win32com.client

        pythoncom.CoInitialize()
        word = None
        doc = None
        try:
            word = win32com.client.DispatchEx("Word.Application")
            word.Visible = False
            doc = word.Documents.Open(os.path.abspath(input_path), ReadOnly=True)
            if callback:
                callback(70, "Exporting PDF...")
            doc.ExportAsFixedFormat(os.path.abspath(output_path), 17)
        finally:
            if doc is not None:
                doc.Close(False)
            if word is not None:
                word.Quit()
            pythoncom.CoUninitialize()
    except Exception as office_error:
        if callback:
            callback(35, "Trying LibreOffice...")
        try:
            _convert_with_soffice(input_path, output_path)
        except Exception as fallback_error:
            raise RuntimeError(
                f"Word export failed: {office_error}. "
                f"LibreOffice fallback failed: {fallback_error}"
            ) from fallback_error

    if callback:
        callback(100, "Done")


def convert_excel_to_pdf(input_path, output_path, callback=None):
    if callback:
        callback(15, "Opening Excel...")

    try:
        import pythoncom
        import win32com.client

        pythoncom.CoInitialize()
        excel = None
        workbook = None
        try:
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            workbook = excel.Workbooks.Open(os.path.abspath(input_path),
                                            ReadOnly=True)
            if callback:
                callback(70, "Exporting PDF...")
            workbook.ExportAsFixedFormat(0, os.path.abspath(output_path))
        finally:
            if workbook is not None:
                workbook.Close(False)
            if excel is not None:
                excel.Quit()
            pythoncom.CoUninitialize()
    except Exception as office_error:
        if callback:
            callback(35, "Trying LibreOffice...")
        try:
            _convert_with_soffice(input_path, output_path)
        except Exception as fallback_error:
            raise RuntimeError(
                f"Excel export failed: {office_error}. "
                f"LibreOffice fallback failed: {fallback_error}"
            ) from fallback_error

    if callback:
        callback(100, "Done")


def _safe_sheet_name(name, used_names):
    clean = re.sub(r"[\[\]\:\*\?\/\\]", "_", name).strip() or "Sheet"
    clean = clean[:31]
    candidate = clean
    index = 2
    while candidate in used_names:
        suffix = f"_{index}"
        candidate = f"{clean[:31 - len(suffix)]}{suffix}"
        index += 1
    used_names.add(candidate)
    return candidate


def convert_pdf_to_excel(input_path, output_path, callback=None):
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise RuntimeError("Install openpyxl first: pip install openpyxl") from exc

    doc = fitz.open(input_path)
    workbook = Workbook()
    workbook.remove(workbook.active)
    used_names = set()

    try:
        total_pages = max(1, len(doc))
        for page_index, page in enumerate(doc):
            if callback:
                pct = int((page_index / total_pages) * 90)
                callback(pct, f"Page {page_index + 1}/{total_pages}")

            sheet_name = _safe_sheet_name(f"Page {page_index + 1}", used_names)
            sheet = workbook.create_sheet(sheet_name)
            row_index = 1
            wrote_content = False

            finder = page.find_tables()
            tables = getattr(finder, "tables", [])
            for table_index, table in enumerate(tables):
                data = table.extract()
                if not data:
                    continue
                if table_index:
                    row_index += 1
                for row in data:
                    for col_index, value in enumerate(row, start=1):
                        sheet.cell(row=row_index, column=col_index,
                                   value=value)
                    row_index += 1
                    wrote_content = True

            if not wrote_content:
                lines = [
                    line.strip()
                    for line in page.get_text("text").splitlines()
                    if line.strip()
                ]
                for line in lines:
                    sheet.cell(row=row_index, column=1, value=line)
                    row_index += 1
                    wrote_content = True

            if not wrote_content:
                sheet.cell(row=1, column=1, value="No table or text found")

            for column_cells in sheet.columns:
                values = [str(cell.value) for cell in column_cells
                          if cell.value is not None]
                width = min(60, max([12] + [len(value) + 2
                                            for value in values]))
                sheet.column_dimensions[column_cells[0].column_letter].width = width
    finally:
        doc.close()

    workbook.save(output_path)
    if callback:
        callback(100, "Done")


def merge_images_to_pdf(image_paths, output_path, callback=None):
    doc = fitz.open()
    try:
        total = max(1, len(image_paths))
        for index, image_path in enumerate(image_paths):
            if callback:
                pct = int((index / total) * 90)
                callback(pct, f"Image {index + 1}/{total}")

            pix = fitz.Pixmap(image_path)
            width = max(1, pix.width)
            height = max(1, pix.height)
            pix = None

            page = doc.new_page(width=width, height=height)
            page.insert_image(page.rect, filename=image_path)

        doc.save(output_path, garbage=4, deflate=True)
    finally:
        doc.close()

    if callback:
        callback(100, "Done")


# ── Colors ───────────────────────────────────────────────────────────
C = {
    "bg":       "#2C3947",
    "surface":  "#34495e",
    "surface2": "#3d5366",
    "border":   "#4a6478",
    "accent":   "#547A95",
    "accent_h": "#6590ab",
    "text":     "#E8EDF2",
    "text2":    "#a8b8c8",
    "green":    "#2ecc71",
    "red":      "#e74c3c",
    "orange":   "#f39c12",
    "white":    "#ffffff",
}


TOOL_CONFIGS = {
    "compress": {
        "label": "Compress",
        "drop": "Drag & drop PDF files here",
        "browse": "Browse PDFs",
        "action": "Compress",
        "filetypes": [("PDF Files", "*.pdf")],
        "extensions": (".pdf",),
        "result_title": "Compression Complete",
    },
    "pdf_to_word": {
        "label": "PDF to Word",
        "drop": "Drag & drop PDF files here",
        "browse": "Browse PDFs",
        "action": "Convert",
        "filetypes": [("PDF Files", "*.pdf")],
        "extensions": (".pdf",),
        "result_title": "Conversion Complete",
    },
    "word_to_pdf": {
        "label": "Word to PDF",
        "drop": "Drag & drop Word files here",
        "browse": "Browse Word",
        "action": "Convert",
        "filetypes": [("Word Files", "*.docx *.doc")],
        "extensions": (".docx", ".doc"),
        "result_title": "Conversion Complete",
    },
    "excel_to_pdf": {
        "label": "Excel to PDF",
        "drop": "Drag & drop Excel files here",
        "browse": "Browse Excel",
        "action": "Convert",
        "filetypes": [("Excel Files", "*.xlsx *.xls *.xlsm")],
        "extensions": (".xlsx", ".xls", ".xlsm"),
        "result_title": "Conversion Complete",
    },
    "pdf_to_excel": {
        "label": "PDF to Excel",
        "drop": "Drag & drop PDF files here",
        "browse": "Browse PDFs",
        "action": "Convert",
        "filetypes": [("PDF Files", "*.pdf")],
        "extensions": (".pdf",),
        "result_title": "Conversion Complete",
    },
    "images_to_pdf": {
        "label": "Images to PDF",
        "drop": "Drag & drop image files here",
        "browse": "Browse Images",
        "action": "Merge",
        "filetypes": [
            ("Image Files", "*.jpg *.jpeg *.png *.bmp *.tif *.tiff *.webp")
        ],
        "extensions": (
            ".jpg", ".jpeg", ".png", ".bmp", ".tif", ".tiff", ".webp"
        ),
        "result_title": "Merge Complete",
    },
}


# ── Application ──────────────────────────────────────────────────────
class PDFCompressorApp(_RootBase):
    def __init__(self):
        super().__init__()
        self.title("PDF Tools  -  by hfzrfa")
        self.geometry("860x780")
        self.minsize(760, 680)
        self.configure(fg_color=C["bg"])
        ctk.set_appearance_mode("dark")

        self.pdf_files: list[str] = []
        self.active_tool = "compress"
        self._settings_file = self._get_settings_file()
        self.last_opened_dir = self._load_last_opened_dir()
        self.output_dir = self.last_opened_dir
        self._is_compressing = False
        self._progress_visible = False

        self._build_ui()
        self._setup_dnd()

    def _get_settings_file(self):
        appdata = os.getenv("APPDATA")
        if appdata:
            return Path(appdata) / "PDF Compressor" / "settings.json"
        return Path.home() / ".pdf_compressor" / "settings.json"

    def _default_output_dir(self):
        for candidate in (Path.home() / "Downloads",
                          Path.home() / "Desktop",
                          Path.home()):
            if candidate.is_dir():
                return str(candidate)
        return str(Path.cwd())

    def _load_last_opened_dir(self):
        default_dir = self._default_output_dir()
        try:
            with self._settings_file.open("r", encoding="utf-8") as fh:
                data = json.load(fh)
            saved_dir = data.get("last_opened_dir")
            if saved_dir and os.path.isdir(saved_dir):
                return saved_dir
        except Exception:
            pass
        return default_dir

    def _save_last_opened_dir(self):
        try:
            self._settings_file.parent.mkdir(parents=True, exist_ok=True)
            with self._settings_file.open("w", encoding="utf-8") as fh:
                json.dump({"last_opened_dir": self.last_opened_dir}, fh)
        except Exception:
            pass

    def _set_last_opened_dir(self, directory):
        if not directory:
            return
        directory = os.path.abspath(os.path.normpath(directory))
        if not os.path.isdir(directory):
            return
        self.last_opened_dir = directory
        self.output_dir = directory
        if hasattr(self, "output_lbl"):
            self.output_lbl.configure(text=directory)
        self._save_last_opened_dir()

    def _tool_config(self):
        return TOOL_CONFIGS[self.active_tool]

    def _on_tool_selected(self, label):
        if self._is_compressing:
            self.tool_selector.set(self._tool_config()["label"])
            return

        for key, config in TOOL_CONFIGS.items():
            if config["label"] == label:
                self.active_tool = key
                break

        if not hasattr(self, "drop_label"):
            return

        self.pdf_files.clear()
        self._refresh_file_list()
        self._reset_progress()
        self._hide_progress_section()
        if hasattr(self, "result_frame") and self.result_frame:
            self.result_frame.pack_forget()
            self.result_frame = None
        self._apply_tool_ui()

    def _apply_tool_ui(self):
        config = self._tool_config()
        self.drop_label.configure(text=config["drop"])
        self.browse_btn.configure(text=config["browse"])
        self.compress_btn.configure(text=config["action"])

        if self.active_tool == "compress":
            if not self.quality_sep.winfo_manager():
                self.quality_sep.pack(fill="x", padx=14)
            if not self.compression_row.winfo_manager():
                self.compression_row.pack(fill="x", padx=14, pady=12)
        else:
            self.quality_sep.pack_forget()
            self.compression_row.pack_forget()

    def _active_extensions(self):
        return self._tool_config()["extensions"]

    # ────────────────────────────────────────────────────────────
    #  UI
    # ────────────────────────────────────────────────────────────
    def _build_ui(self):

        # ── Header ───────────────────────────────────────────────
        header = ctk.CTkFrame(self, fg_color=C["surface"],
                              corner_radius=0, height=52)
        header.pack(fill="x")
        header.pack_propagate(False)

        ctk.CTkLabel(
            header, text="PDF Tools",
            font=ctk.CTkFont("Segoe UI", 18, "bold"),
            text_color=C["white"],
        ).pack(side="left", padx=20)

        ctk.CTkLabel(
            header, text="by hfzrfa",
            font=ctk.CTkFont("Segoe UI", 11, slant="italic"),
            text_color=C["accent_h"],
        ).pack(side="right", padx=20)

        # ── Scrollable body ──────────────────────────────────────
        body = ctk.CTkScrollableFrame(
            self, fg_color="transparent",
            scrollbar_button_color=C["accent"],
            scrollbar_button_hover_color=C["accent_h"],
        )
        body.pack(fill="both", expand=True, padx=0, pady=0)
        self._body = body

        pad = ctk.CTkFrame(body, fg_color="transparent")
        pad.pack(fill="both", expand=True, padx=22, pady=18)

        tool_row = ctk.CTkFrame(pad, fg_color="transparent")
        tool_row.pack(fill="x", pady=(0, 14))
        self.tool_selector = ctk.CTkSegmentedButton(
            tool_row,
            values=[cfg["label"] for cfg in TOOL_CONFIGS.values()],
            fg_color=C["surface"],
            selected_color=C["accent"],
            selected_hover_color=C["accent_h"],
            unselected_color=C["surface2"],
            unselected_hover_color=C["border"],
            text_color=C["text"],
            font=ctk.CTkFont("Segoe UI", 11),
            command=self._on_tool_selected,
        )
        self.tool_selector.pack(fill="x")
        self.tool_selector.set(TOOL_CONFIGS[self.active_tool]["label"])

        # ── 1. Drop Zone ─────────────────────────────────────────
        self.drop_frame = ctk.CTkFrame(
            pad, fg_color=C["surface"], corner_radius=14,
            border_color=C["accent"], border_width=2, height=120)
        self.drop_frame.pack(fill="x")
        self.drop_frame.pack_propagate(False)

        drop_inner = ctk.CTkFrame(self.drop_frame, fg_color="transparent")
        drop_inner.place(relx=0.5, rely=0.5, anchor="center")

        self.drop_label = ctk.CTkLabel(
            drop_inner, text="📂  Drag & drop PDF files here",
            font=ctk.CTkFont("Segoe UI", 14), text_color=C["text2"])
        self.drop_label.pack(pady=(0, 8))

        ctk.CTkButton(
            drop_inner, text="Browse Files", width=130, height=32,
            fg_color=C["accent"], hover_color=C["accent_h"],
            text_color=C["white"], corner_radius=8,
            font=ctk.CTkFont("Segoe UI", 12),
            command=self._browse_files,
        ).pack()
        self.browse_btn = drop_inner.winfo_children()[-1]
        self.drop_label.configure(text=TOOL_CONFIGS[self.active_tool]["drop"])
        self.browse_btn.configure(
            text=TOOL_CONFIGS[self.active_tool]["browse"])

        for w in (self.drop_frame, drop_inner, self.drop_label):
            w.bind("<Button-1>", lambda e: self._browse_files())

        # ── 2. File List ─────────────────────────────────────────
        fh = ctk.CTkFrame(pad, fg_color="transparent", height=26)
        fh.pack(fill="x", pady=(14, 4))
        ctk.CTkLabel(fh, text="Files",
                     font=ctk.CTkFont("Segoe UI", 13, "bold"),
                     text_color=C["text"]).pack(side="left")
        self.file_count_lbl = ctk.CTkLabel(
            fh, text="0 files",
            font=ctk.CTkFont("Segoe UI", 11), text_color=C["text2"])
        self.file_count_lbl.pack(side="right")

        self.file_list = ctk.CTkScrollableFrame(
            pad, fg_color=C["surface"], corner_radius=10, height=110,
            border_color=C["border"], border_width=1,
            scrollbar_button_color=C["accent"],
            scrollbar_button_hover_color=C["accent_h"])
        self.file_list.pack(fill="x")

        self._show_empty_label()

        # ── 3. Settings ──────────────────────────────────────────
        settings = ctk.CTkFrame(pad, fg_color=C["surface"],
                                corner_radius=10,
                                border_color=C["border"], border_width=1)
        settings.pack(fill="x", pady=(14, 0))

        # output dir
        o_row = ctk.CTkFrame(settings, fg_color="transparent")
        o_row.pack(fill="x", padx=14, pady=(12, 0))
        ctk.CTkLabel(o_row, text="Output Folder",
                     font=ctk.CTkFont("Segoe UI", 13, "bold"),
                     text_color=C["text"]).pack(side="left")
        ctk.CTkButton(
            o_row, text="Change", width=68, height=26,
            fg_color=C["surface2"], hover_color=C["border"],
            text_color=C["text"], corner_radius=6,
            font=ctk.CTkFont("Segoe UI", 11),
            command=self._choose_output,
        ).pack(side="right")

        self.output_lbl = ctk.CTkLabel(
            settings, text=self.output_dir, anchor="w",
            font=ctk.CTkFont("Segoe UI", 11), text_color=C["text2"],
            wraplength=660)
        self.output_lbl.pack(fill="x", padx=14, pady=(2, 10))

        self.quality_sep = ctk.CTkFrame(
            settings, fg_color=C["border"], height=1)
        self.quality_sep.pack(fill="x", padx=14)

        # quality
        self.compression_row = ctk.CTkFrame(settings, fg_color="transparent")
        q_row = self.compression_row
        q_row.pack(fill="x", padx=14, pady=12)
        ctk.CTkLabel(q_row, text="Compression",
                     font=ctk.CTkFont("Segoe UI", 13, "bold"),
                     text_color=C["text"]).pack(side="left")

        self.quality_var = ctk.StringVar(value="medium")
        qb = ctk.CTkFrame(q_row, fg_color="transparent")
        qb.pack(side="right")
        for val, lbl in [("low", "Max"), ("medium", "Balanced"),
                         ("high", "Quality")]:
            ctk.CTkRadioButton(
                qb, text=lbl, variable=self.quality_var, value=val,
                font=ctk.CTkFont("Segoe UI", 12),
                fg_color=C["accent"], hover_color=C["accent_h"],
                border_color=C["text2"], text_color=C["text"],
            ).pack(side="left", padx=(12, 0))

        # ── 4. Progress Section ──────────────────────────────────
        self.prog_card = ctk.CTkFrame(
            pad, fg_color=C["surface"], corner_radius=10,
            border_color=C["border"], border_width=1)
        prog_card = self.prog_card

        prog_inner = ctk.CTkFrame(prog_card, fg_color="transparent")
        prog_inner.pack(fill="x", padx=14, pady=14)

        # top row: label + percentage
        prog_top = ctk.CTkFrame(prog_inner, fg_color="transparent")
        prog_top.pack(fill="x", pady=(0, 6))

        self.prog_title = ctk.CTkLabel(
            prog_top, text="Progress",
            font=ctk.CTkFont("Segoe UI", 13, "bold"),
            text_color=C["text"])
        self.prog_title.pack(side="left")

        self.pct_label = ctk.CTkLabel(
            prog_top, text="0%",
            font=ctk.CTkFont("Segoe UI", 13, "bold"),
            text_color=C["accent"])
        self.pct_label.pack(side="right")

        # progress bar
        self.progress_bar = ctk.CTkProgressBar(
            prog_inner, height=14, corner_radius=7,
            fg_color=C["surface2"], progress_color=C["accent"])
        self.progress_bar.pack(fill="x")
        self.progress_bar.set(0)

        # status text
        self.status_lbl = ctk.CTkLabel(
            prog_inner, text="Waiting...", anchor="w",
            font=ctk.CTkFont("Segoe UI", 11), text_color=C["text2"])
        self.status_lbl.pack(fill="x", pady=(6, 0))

        # ── 5. Action Buttons ────────────────────────────────────
        self.btn_row = ctk.CTkFrame(pad, fg_color="transparent")
        btn_row = self.btn_row
        btn_row.pack(fill="x", pady=(16, 0))

        ctk.CTkButton(
            btn_row, text="Clear All", width=110, height=40,
            fg_color=C["surface"], hover_color=C["surface2"],
            text_color=C["text"], corner_radius=8,
            border_color=C["border"], border_width=1,
            font=ctk.CTkFont("Segoe UI", 13),
            command=self._clear_files,
        ).pack(side="left")

        self.compress_btn = ctk.CTkButton(
            btn_row, text="Compress", width=180, height=40,
            fg_color=C["accent"], hover_color=C["accent_h"],
            text_color=C["white"], corner_radius=8,
            font=ctk.CTkFont("Segoe UI", 14, "bold"),
            command=self._start_compress)
        self.compress_btn.pack(side="right")

        # ── 6. Results (hidden until compression finishes) ───────
        self._result_parent = pad
        self.result_frame = None

        # ── Footer ───────────────────────────────────────────────
        footer = ctk.CTkFrame(self, fg_color=C["surface"],
                              corner_radius=0, height=30)
        footer.pack(fill="x", side="bottom")
        footer.pack_propagate(False)
        ctk.CTkLabel(
            footer, text="PDF Tools  -  by hfzrfa",
            font=ctk.CTkFont("Segoe UI", 10),
            text_color=C["text2"],
        ).pack(expand=True)

    # ────────────────────────────────────────────────────────────
    #  Drag & Drop
    # ────────────────────────────────────────────────────────────
    def _setup_dnd(self):
        if not DND_AVAILABLE:
            return
        try:
            self.drop_target_register(DND_FILES)
            self.dnd_bind("<<Drop>>", self._on_drop)
            self.dnd_bind("<<DragEnter>>", self._on_drag_enter)
            self.dnd_bind("<<DragLeave>>", self._on_drag_leave)
        except Exception:
            pass

    def _on_drag_enter(self, event=None):
        self.drop_frame.configure(border_color=C["green"])
        self.drop_label.configure(text="✅  Release to add files!")
        return event.action if event else None

    def _on_drag_leave(self, event=None):
        self.drop_frame.configure(border_color=C["accent"])
        self.drop_label.configure(text="📂  Drag & drop PDF files here")
        self.drop_label.configure(text=self._tool_config()["drop"])
        return event.action if event else None

    def _on_drop(self, event):
        self._on_drag_leave()
        raw = event.data
        files = []
        if "{" in raw:
            import re
            files = re.findall(r"\{([^}]+)\}", raw)
            rest = re.sub(r"\{[^}]+\}", "", raw).strip()
            if rest:
                files.extend(rest.split())
        else:
            files = raw.split()
        added = 0
        last_pdf_dir = None
        active_exts = self._active_extensions()
        for f in files:
            f = f.strip()
            if f.lower().endswith(active_exts) and os.path.isfile(f):
                last_pdf_dir = os.path.dirname(os.path.abspath(f))
                if f not in self.pdf_files:
                    self.pdf_files.append(f)
                    added += 1
        if last_pdf_dir:
            self._set_last_opened_dir(last_pdf_dir)
        self._refresh_file_list()
        if added:
            self.status_lbl.configure(
                text=f"Added {added} file(s) via drag & drop")
        return event.action

    # ────────────────────────────────────────────────────────────
    #  File Operations
    # ────────────────────────────────────────────────────────────
    def _browse_files(self):
        if self._is_compressing:
            return
        paths = filedialog.askopenfilenames(
            title="Select Files",
            initialdir=self.last_opened_dir,
            filetypes=self._tool_config()["filetypes"])
        added = 0
        for p in paths:
            if p not in self.pdf_files:
                self.pdf_files.append(p)
                added += 1
        if paths:
            self._set_last_opened_dir(os.path.dirname(paths[-1]))
        self._refresh_file_list()
        if added:
            self.status_lbl.configure(text=f"Added {added} file(s)")

    def _choose_output(self):
        d = filedialog.askdirectory(title="Choose Output Folder",
                                    initialdir=self.last_opened_dir)
        if d:
            self._set_last_opened_dir(d)

    def _remove_file(self, path):
        if path in self.pdf_files:
            self.pdf_files.remove(path)
        self._refresh_file_list()

    def _show_progress_section(self):
        if self._progress_visible:
            return
        self.prog_card.pack(fill="x", pady=(14, 0), before=self.btn_row)
        self._progress_visible = True

    def _hide_progress_section(self):
        if not self._progress_visible:
            return
        self.prog_card.pack_forget()
        self._progress_visible = False

    def _reset_progress(self, status="Waiting..."):
        self.progress_bar.set(0)
        self.progress_bar.configure(progress_color=C["accent"])
        self.pct_label.configure(text="0%")
        self.status_lbl.configure(text=status)

    def _clear_files(self):
        self.pdf_files.clear()
        self._refresh_file_list()
        self._reset_progress()
        self._hide_progress_section()
        if self.result_frame:
            self.result_frame.pack_forget()
            self.result_frame = None

    def _show_empty_label(self):
        ctk.CTkLabel(
            self.file_list, text="No files added yet",
            font=ctk.CTkFont("Segoe UI", 12),
            text_color=C["text2"]).pack(pady=20)

    def _refresh_file_list(self):
        for w in self.file_list.winfo_children():
            w.destroy()
        if not self.pdf_files:
            self._show_empty_label()
            self.file_count_lbl.configure(text="0 files")
            return

        self.file_count_lbl.configure(text=f"{len(self.pdf_files)} file(s)")
        for fp in self.pdf_files:
            row = ctk.CTkFrame(self.file_list, fg_color=C["surface2"],
                               corner_radius=6, height=34)
            row.pack(fill="x", pady=2, padx=2)
            row.pack_propagate(False)

            ctk.CTkLabel(
                row, text=os.path.basename(fp), anchor="w",
                font=ctk.CTkFont("Segoe UI", 12),
                text_color=C["text"]).pack(side="left", padx=(12, 6))
            ctk.CTkLabel(
                row, text=fmt(os.path.getsize(fp)),
                font=ctk.CTkFont("Segoe UI", 10),
                text_color=C["text2"]).pack(side="left")

            _fp = fp
            ctk.CTkButton(
                row, text="✕", width=26, height=26,
                fg_color="transparent", hover_color=C["red"],
                text_color=C["text2"], corner_radius=4,
                font=ctk.CTkFont(size=12),
                command=lambda p=_fp: self._remove_file(p),
            ).pack(side="right", padx=4)

    # ────────────────────────────────────────────────────────────
    #  Compression
    # ────────────────────────────────────────────────────────────
    def _make_output_path(self, input_path, tool_key):
        base = Path(input_path).stem
        if tool_key == "compress":
            return os.path.join(self.output_dir, f"{base}_compressed.pdf")
        if tool_key == "pdf_to_word":
            return unique_output_path(self.output_dir, f"{base}_converted",
                                      ".docx")
        if tool_key == "word_to_pdf":
            return unique_output_path(self.output_dir, f"{base}_converted",
                                      ".pdf")
        if tool_key == "excel_to_pdf":
            return unique_output_path(self.output_dir, f"{base}_converted",
                                      ".pdf")
        if tool_key == "pdf_to_excel":
            return unique_output_path(self.output_dir, f"{base}_converted",
                                      ".xlsx")
        raise RuntimeError("Unsupported tool selected.")

    def _start_compress(self):
        if self._is_compressing:
            return
        if not self.pdf_files:
            messagebox.showwarning("No Files",
                                   "Please add files first.")
            return
        if not os.path.isdir(self.output_dir):
            messagebox.showerror("Invalid Folder",
                                 "Output folder does not exist.")
            return

        self._is_compressing = True
        self.compress_btn.configure(state="disabled", text="Working...")
        self._show_progress_section()
        self._reset_progress("Starting...")
        if self.result_frame:
            self.result_frame.pack_forget()
            self.result_frame = None
        threading.Thread(target=self._compress_thread, daemon=True).start()

    def _compress_thread(self):
        active_tool = self.active_tool
        quality = self.quality_var.get()
        files = list(self.pdf_files)
        results = []
        total = len(files)
        success_count = 0
        fail_count = 0

        if active_tool == "images_to_pdf":
            out_path = unique_output_path(self.output_dir, "merged_images", ".pdf")

            def merge_cb(pct, msg):
                self.after(0, self._update_progress, pct, msg)

            try:
                original_size = sum(os.path.getsize(fp) for fp in files)
                merge_images_to_pdf(files, out_path, merge_cb)
                output_size = os.path.getsize(out_path)
                results.append({
                    "name": f"{len(files)} image(s)",
                    "input": original_size,
                    "output": output_size,
                    "ok": True,
                    "err": "",
                    "out_path": out_path,
                    "kind": active_tool,
                })
                success_count += 1
            except Exception as e:
                results.append({
                    "name": "Images to PDF",
                    "input": 0,
                    "output": 0,
                    "ok": False,
                    "err": str(e),
                    "out_path": "",
                    "kind": active_tool,
                })
                fail_count += 1

            self.after(0, self._compression_done,
                       results, success_count, fail_count)
            return

        for i, fp in enumerate(files):
            name = os.path.basename(fp)
            base, ext = os.path.splitext(name)
            out_path = self._make_output_path(fp, active_tool)

            if os.path.normpath(out_path) == os.path.normpath(fp):
                out_name = f"{base}_compressed_out{ext}"
                out_path = os.path.join(self.output_dir, out_name)

            def cb(pct, msg, _i=i, _total=total, _name=name):
                overall = int(((_i + pct / 100) / _total) * 100)
                self.after(0, self._update_progress, overall,
                           f"[{_i+1}/{_total}]  {_name}  —  {msg}")

            try:
                orig = os.path.getsize(fp)
                if active_tool == "compress":
                    orig, output_size = compress_single_pdf(
                        fp, out_path, quality, cb)
                elif active_tool == "pdf_to_word":
                    convert_pdf_to_word(fp, out_path, cb)
                    output_size = os.path.getsize(out_path)
                elif active_tool == "word_to_pdf":
                    convert_word_to_pdf(fp, out_path, cb)
                    output_size = os.path.getsize(out_path)
                elif active_tool == "excel_to_pdf":
                    convert_excel_to_pdf(fp, out_path, cb)
                    output_size = os.path.getsize(out_path)
                elif active_tool == "pdf_to_excel":
                    convert_pdf_to_excel(fp, out_path, cb)
                    output_size = os.path.getsize(out_path)
                else:
                    raise RuntimeError("Unsupported tool selected.")

                results.append({
                    "name": name,
                    "input": orig,
                    "output": output_size,
                    "ok": True,
                    "err": "",
                    "out_path": out_path,
                    "kind": active_tool,
                })
                success_count += 1
            except Exception as e:
                results.append({
                    "name": name,
                    "input": 0,
                    "output": 0,
                    "ok": False,
                    "err": str(e),
                    "out_path": "",
                    "kind": active_tool,
                })
                fail_count += 1

        self.after(0, self._compression_done,
                   results, success_count, fail_count)

    def _update_progress(self, pct, msg):
        clamped = max(0, min(100, pct))
        self.progress_bar.set(clamped / 100)
        self.pct_label.configure(text=f"{clamped}%")
        self.status_lbl.configure(text=msg)

    def _render_tool_results(self, results, ok_n, fail_n):
        if self.result_frame:
            self.result_frame.pack_forget()

        self.result_frame = ctk.CTkFrame(
            self._result_parent, fg_color=C["surface"],
            corner_radius=10, border_color=C["border"], border_width=1)
        self.result_frame.pack(fill="x", pady=(14, 0))

        ctk.CTkLabel(
            self.result_frame, text="Results",
            font=ctk.CTkFont("Segoe UI", 13, "bold"),
            text_color=C["text"]).pack(anchor="w", padx=14, pady=(12, 6))

        total_input = 0
        total_output = 0

        for item in results:
            row = ctk.CTkFrame(self.result_frame, fg_color=C["surface2"],
                               corner_radius=6, height=40)
            row.pack(fill="x", padx=10, pady=2)
            row.pack_propagate(False)

            if item["ok"]:
                total_input += item["input"]
                total_output += item["output"]
                if item["kind"] == "compress" and item["input"] > 0:
                    saved_pct = (1 - item["output"] / item["input"]) * 100
                    detail = (f"{fmt(item['input'])} -> "
                              f"{fmt(item['output'])}   {saved_pct:.1f}%")
                    clr = C["green"] if saved_pct > 20 else C["text2"]
                else:
                    detail = f"{fmt(item['input'])} -> {fmt(item['output'])}"
                    clr = C["green"]

                ctk.CTkLabel(
                    row, text=f"OK  {item['name']}", anchor="w",
                    font=ctk.CTkFont("Segoe UI", 11),
                    text_color=C["text"]).pack(side="left", padx=10)
                ctk.CTkLabel(
                    row, text=detail,
                    font=ctk.CTkFont("Segoe UI", 11, "bold"),
                    text_color=clr).pack(side="right", padx=10)
            else:
                ctk.CTkLabel(
                    row, text=f"FAILED  {item['name']}: {item['err']}",
                    anchor="w", font=ctk.CTkFont("Segoe UI", 11),
                    text_color=C["red"]).pack(side="left", padx=10)

        sep = ctk.CTkFrame(self.result_frame, fg_color=C["border"], height=1)
        sep.pack(fill="x", padx=14, pady=(8, 0))

        summary = ctk.CTkFrame(self.result_frame, fg_color="transparent")
        summary.pack(fill="x", padx=14, pady=(6, 12))

        if self.active_tool == "compress" and total_input > 0:
            saved = total_input - total_output
            saved_pct = (saved / total_input) * 100
            summary_text = f"Total saved: {fmt(saved)} ({saved_pct:.1f}%)"
        elif ok_n:
            summary_text = f"Created: {ok_n} output file(s)"
        else:
            summary_text = "No output created"

        ctk.CTkLabel(
            summary, text=summary_text,
            font=ctk.CTkFont("Segoe UI", 12, "bold"),
            text_color=C["text"]).pack(side="left")
        ctk.CTkLabel(
            summary, text=f"Output: {self.output_dir}",
            font=ctk.CTkFont("Segoe UI", 10),
            text_color=C["text2"]).pack(side="right")

        title = self._tool_config()["result_title"]
        if fail_n == 0:
            self.status_lbl.configure(text="All tasks completed!")
            self.progress_bar.configure(progress_color=C["green"])
            messagebox.showinfo(
                title,
                f"{ok_n} task(s) completed successfully.\n\n"
                f"Output folder:\n{self.output_dir}")
        elif ok_n == 0:
            self.status_lbl.configure(text="All tasks failed!")
            self.progress_bar.configure(progress_color=C["red"])
            err_details = "\n".join(
                f"- {item['name']}: {item['err']}"
                for item in results if not item["ok"]
            )
            messagebox.showerror(
                "Task Failed",
                f"{fail_n} task(s) failed.\n\n{err_details}")
        else:
            self.status_lbl.configure(
                text=f"{ok_n} success, {fail_n} failed")
            self.progress_bar.configure(progress_color=C["orange"])
            messagebox.showwarning(
                "Partial Success",
                f"{ok_n} task(s) completed\n"
                f"{fail_n} task(s) failed\n\n"
                f"Output folder:\n{self.output_dir}")

    def _compression_done(self, results, ok_n, fail_n):
        self._is_compressing = False
        self.compress_btn.configure(
            state="normal", text=self._tool_config()["action"])
        self.progress_bar.set(1)
        self.pct_label.configure(text="100%")
        self._render_tool_results(results, ok_n, fail_n)
        self.after(3000, lambda: self.progress_bar.configure(
            progress_color=C["accent"]))
        return

        # ── Build result panel ───────────────────────────────────
        if self.result_frame:
            self.result_frame.pack_forget()

        self.result_frame = ctk.CTkFrame(
            self._result_parent, fg_color=C["surface"],
            corner_radius=10, border_color=C["border"], border_width=1)
        self.result_frame.pack(fill="x", pady=(14, 0))

        ctk.CTkLabel(
            self.result_frame, text="Results",
            font=ctk.CTkFont("Segoe UI", 13, "bold"),
            text_color=C["text"]).pack(anchor="w", padx=14, pady=(12, 6))

        total_orig = 0
        total_comp = 0

        for name, orig, comp, ratio, ok, err in results:
            row = ctk.CTkFrame(self.result_frame, fg_color=C["surface2"],
                               corner_radius=6, height=36)
            row.pack(fill="x", padx=10, pady=2)
            row.pack_propagate(False)

            if ok:
                total_orig += orig
                total_comp += comp
                clr = (C["green"] if ratio > 20
                       else C["orange"] if ratio > 5
                       else C["text2"])

                ctk.CTkLabel(
                    row, text=f"✅  {name}", anchor="w",
                    font=ctk.CTkFont("Segoe UI", 11),
                    text_color=C["text"]).pack(side="left", padx=10)
                ctk.CTkLabel(
                    row,
                    text=f"{fmt(orig)} → {fmt(comp)}   −{ratio:.1f}%",
                    font=ctk.CTkFont("Segoe UI", 11, "bold"),
                    text_color=clr).pack(side="right", padx=10)
            else:
                ctk.CTkLabel(
                    row, text=f"❌  {name}:  {err}", anchor="w",
                    font=ctk.CTkFont("Segoe UI", 11),
                    text_color=C["red"]).pack(side="left", padx=10)

        # ── Summary row ──────────────────────────────────────────
        if total_orig > 0:
            saved = total_orig - total_comp
            saved_pct = (saved / total_orig) * 100

            sep = ctk.CTkFrame(self.result_frame, fg_color=C["border"],
                               height=1)
            sep.pack(fill="x", padx=14, pady=(8, 0))

            summary = ctk.CTkFrame(self.result_frame, fg_color="transparent")
            summary.pack(fill="x", padx=14, pady=(6, 12))

            ctk.CTkLabel(
                summary, text="Total saved:",
                font=ctk.CTkFont("Segoe UI", 12),
                text_color=C["text"]).pack(side="left")
            ctk.CTkLabel(
                summary,
                text=f"{fmt(saved)}  ({saved_pct:.1f}%)",
                font=ctk.CTkFont("Segoe UI", 13, "bold"),
                text_color=C["green"]).pack(side="left", padx=(8, 0))

            ctk.CTkLabel(
                summary, text=f"Output:  {self.output_dir}",
                font=ctk.CTkFont("Segoe UI", 10),
                text_color=C["text2"]).pack(side="right")
        else:
            ctk.CTkFrame(self.result_frame, fg_color="transparent",
                         height=8).pack()

        # ── Status & notification popup ──────────────────────────
        if fail_n == 0:
            self.status_lbl.configure(text="✅  All files compressed!")
            self.progress_bar.configure(progress_color=C["green"])
            messagebox.showinfo(
                "Compression Complete",
                f"✅  {ok_n} file(s) compressed successfully!\n\n"
                f"Saved: {fmt(total_orig - total_comp)}  "
                f"({(total_orig - total_comp) / total_orig * 100:.1f}%)\n\n"
                f"Output folder:\n{self.output_dir}")
        elif ok_n == 0:
            self.status_lbl.configure(text="❌  All files failed!")
            self.progress_bar.configure(progress_color=C["red"])
            # Collect error details for popup
            err_details = "\n".join(
                f"• {name}: {err}"
                for name, _, _, _, ok, err in results if not ok
            )
            messagebox.showerror(
                "Compression Failed",
                f"❌  {fail_n} file(s) failed to compress.\n\n"
                f"{err_details}")
        else:
            self.status_lbl.configure(
                text=f"⚠️  {ok_n} success, {fail_n} failed")
            self.progress_bar.configure(progress_color=C["orange"])
            messagebox.showwarning(
                "Partial Success",
                f"✅  {ok_n} file(s) compressed\n"
                f"❌  {fail_n} file(s) failed\n\n"
                f"Output folder:\n{self.output_dir}")

        # Reset progress bar color after a delay
        self.after(3000, lambda: self.progress_bar.configure(
            progress_color=C["accent"]))


# ── Entry Point ──────────────────────────────────────────────────────
if __name__ == "__main__":
    app = PDFCompressorApp()
    app.mainloop()
